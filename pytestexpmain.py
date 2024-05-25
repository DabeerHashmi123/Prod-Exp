import sys
import os
from PyQt5.QtWidgets import QApplication, QAbstractScrollArea,QMainWindow, QAction, QToolBar, QMessageBox, QStatusBar, QTableWidget, QTableWidgetItem, QFileDialog, QInputDialog, QWidget, QVBoxLayout, QPushButton, QLabel, QHBoxLayout, QLineEdit,QListWidgetItem,QListWidget,QDialog,QAbstractItemView,QDateEdit
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QTimer, QDate
from win10toast import ToastNotifier
import sqlite3
from openpyxl import load_workbook
from inspect import getsourcefile
from fuzzywuzzy import process as fw_process
from datetime import datetime
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QLineEdit
import module_locator

class NearExpiryDialog(QDialog):
    def __init__(self, near_expiry_products):
        super().__init__()
        print("init near expiry")
        self.setWindowTitle("Near Expiry Products")
        self.resize(800, 600)  # Set the size of the dialog window

        layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)
        self.setLayout(layout)

        self.populate_table(near_expiry_products)

    def populate_table(self, products):
        print("populate table")
        self.table_widget.setColumnCount(6)  # Set the number of columns

        # Set the header labels for each column
        self.table_widget.setHorizontalHeaderLabels(["ID", "Barcode", "Product Name", "Quantity","Vendor", "Expiry Date"])

        # Set the number of rows based on the number of products
        self.table_widget.setRowCount(len(products))

        for row, product in enumerate(products):
            # Populate each cell in the table
            for col, item in enumerate(product):
                # Create QTableWidgetItem for each item
                table_item = QTableWidgetItem(str(item))
                # Set alignment for each cell
                if col == 0:  # ID column
                    table_item.setTextAlignment(Qt.AlignCenter)  # Center alignment
                else:
                    table_item.setTextAlignment(Qt.AlignLeft)  # Left alignment
                self.table_widget.setItem(row, col, table_item)

        # Resize the columns to fit the contents
        self.table_widget.resizeColumnsToContents()

class MainWindow(QMainWindow):
    def __init__(self, pth):
        print("init")
        super().__init__()
        my_path = pth
        
        dir_path = my_path
        db_path = dir_path + "\\P-E-N-S-main\\new_database1.db"
        print(db_path)
        if not os.path.exists(db_path):
            try:
                self.conn = sqlite3.connect(db_path)
                self.cur = self.conn.cursor()
                self.cur.execute('''CREATE TABLE IF NOT EXISTS products (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    barcode TEXT,
                                    product_name TEXT,
                                    quantity TEXT,
                                    vendor TEXT,
                                    expiry_date DATE
                                    )''')
                self.conn.commit()
            except Exception as e:
                print(f"Error in SQLite: {e}")
                print(db_path)
                sys.exit(1)
        else:
            try:
                self.conn = sqlite3.connect(db_path)
                self.cur = self.conn.cursor()
            except Exception as e:
                print(f"Error in SQLite: {e}")
                print(db_path)
                sys.exit(1)

        try:
            self.toaster = ToastNotifier()
        except Exception as e:
            print(f"Error in win10toast: {e}")
            sys.exit(1)
        
        self.near_expiry_products = []

        self.setWindowTitle("Product Expiry Notification System")
        self.setWindowIcon(QIcon(dir_path+"\\P-E-N-S-main\\notification-icon-bell-alarm2.ico"))
        self.setMinimumSize(1024, 768)

        toolbar = QToolBar()
        self.addToolBar(toolbar)

        import_data_action = QAction(QIcon(dir_path+"\\P-E-N-S-main\\add-product.png"), "Import Data", self)
        import_data_action.triggered.connect(self.import_data)
        toolbar.addAction(import_data_action)

        search_action = QAction(QIcon(dir_path+"\\P-E-N-S-main\\s1.png"), "Search", self)
        search_action.triggered.connect(self.search)
        toolbar.addAction(search_action)

        refresh_action = QAction(QIcon(dir_path+"\\P-E-N-S-main\\r3.png"), "Refresh", self)
        refresh_action.triggered.connect(self.refresh_data)  # Change this to self.refresh_data
        toolbar.addAction(refresh_action)
        
        near_expiry_action = QAction(QIcon(dir_path+"\\P-E-N-S-main\\download.png"), "Near Expiry Products", self)
        near_expiry_action.triggered.connect(self.show_near_expiry_list)
        toolbar.addAction(near_expiry_action)

        delete_action = QAction(QIcon(dir_path+"\\P-E-N-S-main\\d1.png"), "Delete", self)
        delete_action.triggered.connect(self.delete_product)
        toolbar.addAction(delete_action)

        statusbar = QStatusBar()
        self.setStatusBar(statusbar)

        self.tableWidget = QTableWidget()
        self.setCentralWidget(self.tableWidget)
        self.tableWidget.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.tableWidget.setAlternatingRowColors(True)
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels(("ID", "Barcode", "Product Name", "Quantity", "Vendor", "Expiry Date"))
        print("loading data")
        self.load_data()
        self.tableWidget.resizeColumnsToContents()

       # Setup timer for notifications
        self.notification_timer = QTimer(self)
        self.notification_timer.timeout.connect(self.check_expiry)

       # Initial notification after 2 seconds
        QTimer.singleShot(2000, self.check_expiry)

       # Start the timer to check every 30 minutes
        self.notification_timer.start(30 * 60 * 1000)  # 30 minutes in milliseconds

       # Set inline CSS styles for QMainWindow
        self.setStyleSheet("""
    QMainWindow {
        background-color: #f0f0f0; /* Light grey */
    }
    QToolBar {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6A0DAD, stop:0.5 #A74DF6, stop:0.9 #FAC8FD, stop:1 #A74DF6) !important;
    color: #fff; /* Text color */
    }

    QTableWidget {
        background-color: #fff; /* White */
        border-radius: 10px;
    }

    QTableWidget QHeaderView::section {
        background-color: #3c486a; /* Dark blue-gray */
    color: #fff; /* Text color */
    padding: 8px;
    border-radius: 0;
    }

    QTableWidget::item {
        padding: 10px;
    }

    QAction {
        background-color: #4CAF50; /* Green */
    border: none;
    color: white;
    padding: 12px 24px;
    text-align: center;
    text-decoration: none;
    display: inline-block;
    font-size: 16px;
    margin: 4px 2px;
    transition: transform 0.3s ease; /* Add transition for smooth transform */
    border-radius: 12px;
    }

    QAction:hover {
        background-color: #45a049; /* Dark Green */
    transform: scale(1.2); /* Scale up on hover */
    }

    QAction:active {
    transform: scale(0.9); /* Scale down when clicked */
    }                       

    QLabel {
        color: #333;
    }
""")

    def import_data(self) -> None:
     """Prompt for file selection and import data from Excel file."""
     file_path, _ = QFileDialog.getOpenFileName(
        self, "Select File", "", "Excel Files (*.xlsx *.xls)")

     if file_path:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        # Define expected column names
        expected_columns = {'Barcode', 'Product Name', 'Quantity', 'Vendor'}

        # Extract actual column names from the header row
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        actual_columns = set(header_row)

        # Match actual column names to expected column names
        column_mapping = {}
        for expected_column in expected_columns:
            matched_column, _ = fw_process.extractOne(expected_column, actual_columns)
            if matched_column:
                column_mapping[expected_column] = header_row.index(matched_column)
            else:
                QMessageBox.warning(
                    self, "Warning", f"Could not find a match for '{expected_column}' column.")

        # Check if all expected columns have been matched
        if len(column_mapping) < len(expected_columns):
            QMessageBox.critical(
                self, "Error", "Could not find matches for all required columns "
                                "in the Excel file.")
            return  # Exit the function if not all columns are matched

        # Proceed with data extraction and import using column_mapping
        try:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                barcode = row[column_mapping['Barcode']]
                product_name = row[column_mapping['Product Name']]
                quantity = row[column_mapping['Quantity']]
                vendor = row[column_mapping['Vendor']]
                expiry_date = None  # Set expiry_date to None

                self.cur.execute(
                    "INSERT INTO products (barcode, product_name, quantity, vendor, expiry_date) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (barcode, product_name, quantity, vendor, expiry_date))
            self.conn.commit()
            self.load_data()  # Load data after all rows have been inserted
            QMessageBox.information(self, "Success", "Data imported successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")


    def search(self) -> None:
     """Enhanced real-time search for products."""
     search_text, ok_pressed = QInputDialog.getText(self, "Search", "Enter text to search:")
     if ok_pressed:
        search_text = search_text.strip().lower()  # Convert search text to lowercase and remove leading/trailing spaces
        if search_text:
            self.tableWidget.clearSelection()  # Clear previous selections

            # Clear previous highlights
            for row in range(self.tableWidget.rowCount()):
                for column in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, column)
                    if item:
                        item.setBackground(Qt.white)  # Reset background to white (or default color)

            found = False
            matches = 0
            for row in range(self.tableWidget.rowCount()):
                for column in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, column)
                    if item and search_text in item.text().lower():
                        item.setBackground(Qt.yellow)  # Highlight the cell
                        found = True
                        matches += 1

            if found:
                QMessageBox.information(self, "Search Result", f"{matches} matching product(s) found.")
            else:
                QMessageBox.information(self, "Search Result", "No matching product found.")
        else:
            QMessageBox.information(self, "Search Result", "Please enter a search term.")

    def check_expiry(self):
        """Check products with near expiry dates and store them."""
        try:
            current_date = QDate.currentDate()
            expiry_threshold = current_date.addDays(7)

            self.cur.execute("SELECT * FROM products WHERE expiry_date <= ? AND expiry_date != '1752-09-14'", 
                         (expiry_threshold.toString("yyyy-MM-dd"),))
            self.near_expiry_products = self.cur.fetchall()

            if self.near_expiry_products:
                QMessageBox.information(self, "Near Expiry Alert", "You have some products near expiry.")

        except Exception as e:
            print(f"Error checking expiry: {e}")

    def delete_product(self) -> None:
     """Delete selected products from the table and database."""
     selected_rows = list(set(index.row() for index in self.tableWidget.selectedIndexes()))
     if selected_rows:
        ids_to_delete = []
        for row in selected_rows:
            item = self.tableWidget.item(row, 0)  # Assuming the ID is in the first column
            if item:
                ids_to_delete.append(item.text())

        if ids_to_delete:
            confirmation = QMessageBox.question(
                self, "Confirm Delete",
                f"Are you sure you want to delete {len(ids_to_delete)} selected products?",
                QMessageBox.Yes | QMessageBox.No
            )
            if confirmation == QMessageBox.Yes:
                try:
                    self.cur.executemany("DELETE FROM products WHERE id = ?", [(id,) for id in ids_to_delete])
                    self.conn.commit()
                    self.load_data()
                    QMessageBox.information(self, "Success", "Selected products deleted successfully.")
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
        else:
            QMessageBox.warning(self, "Warning", "No valid product IDs found for deletion.")
     else:
        QMessageBox.warning(self, "Warning", "Please select products to delete.")

    def load_data(self) -> None:
     """Load non-empty data from the database into the table widget."""
     try:
        self.cur.execute("SELECT * FROM products WHERE barcode IS NOT NULL AND product_name IS NOT NULL AND quantity IS NOT NULL AND vendor IS NOT NULL")
        data = self.cur.fetchall()
        self.tableWidget.setRowCount(0)
        for row_number, row_data in enumerate(data):
            self.tableWidget.insertRow(row_number)
            for column_number, item_data in enumerate(row_data):
                if column_number == 5:  # Assuming the expiry_date column index is 5
                    if item_data is None or item_data == "1752-09-14":  # Check if date is default or None
                        item_data = ""
                    date_obj = QDate.fromString(item_data, "yyyy-MM-dd")
                    date_edit = QDateEdit(date_obj)
                    date_edit.setDisplayFormat("yyyy-MM-dd")
                    date_edit.setCalendarPopup(True)
                    date_edit.dateChanged.connect(lambda date, row=row_number: self.update_expiry_date(date, row))
                    self.tableWidget.setCellWidget(row_number, column_number, date_edit)
                else:
                    item = QTableWidgetItem(str(item_data))
                    self.tableWidget.setItem(row_number, column_number, item)
     except Exception as e:
        QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

    def update_expiry_date(self, new_date, row):
     """Update the expiry date in the database when changed in the table."""
     try:
        item = self.tableWidget.item(row, 0)  # Assuming the ID column index is 0
        product_id = int(item.text())
        self.cur.execute("UPDATE products SET expiry_date = ? WHERE id = ?", (new_date.toString("yyyy-MM-dd"), product_id))
        self.conn.commit()
     except Exception as e:
        QMessageBox.critical(self, "Error", f"An error occurred while updating expiry date: {str(e)}")

    def refresh_data(self) -> None:
        """Clear the table and reload data from the database."""
        confirmation = QMessageBox.question(self, "Confirmation", "Are you sure you want to refresh and clear all data?",
                                             QMessageBox.Yes | QMessageBox.No)
        if confirmation == QMessageBox.Yes:
            try:
                self.cur.execute("DELETE FROM products")
                self.conn.commit()
                self.load_data()
                QMessageBox.information(self, "Success", "Database cleared successfully.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
    
    def show_near_expiry_list(self):
       try:
        current_date = QDate.currentDate()
        expiry_threshold = current_date.addDays(15)

        self.cur.execute("SELECT * FROM products WHERE expiry_date <= ?", (expiry_threshold.toString("yyyy-MM-dd"),))
        data = self.cur.fetchall()

        # Create a new dialog to display near expiry products
        dialog = NearExpiryDialog(data)
        dialog.exec_()
        
       except Exception as e:
        print(f"Error fetching near expiry products: {e}")

if __name__ == "__main__":
    app = QApplication([])
    file_path = module_locator.module_path()
    main_window = MainWindow(file_path)
    main_window.show()
    try:
        sys.exit(app.exec())
    except Exception as e:
        print(f"An error occurred: {e}")
